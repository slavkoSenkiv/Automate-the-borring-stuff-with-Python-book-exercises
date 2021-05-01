import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
"""boldFont = Font(bold=True)
sheet.min_row.font = boldFont
sheet.min_column.font = boldFont"""

userNumber = 3 # int(input('Enter number...'))

# Column A
for firstCol in range(1, userNumber + 1):
    sheet.cell(row=firstCol + 1, column=1).value = firstCol
# Row 1
for firstRow in range(1, userNumber + 1):
    sheet.cell(row=1, column=firstRow + 1).value = firstRow

for rowsInColumn in range(2, userNumber + 2):
    for cellsInRow in range(2, userNumber + 2):
        sheet.cell(row=rowsInColumn, column=cellsInRow).value = (rowsInColumn - 1) * (cellsInRow - 1)

wb.save('table.xlsx')
