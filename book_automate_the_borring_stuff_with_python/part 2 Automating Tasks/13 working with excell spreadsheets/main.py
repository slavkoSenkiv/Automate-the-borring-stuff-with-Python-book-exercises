"""
import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(type(wb))


import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.sheetnames) # the workbook sheet names

sheet = wb['Sheet3'] # get the sheet from the work book
print(sheet)

print(type(sheet))

print(sheet.title) # get the sheets title as a string

anotherSheet = wb.active # get the active sheet
print(anotherSheet)

# Getting Cells from the Sheets

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1'] # get the sheet from the workbook
print(sheet['A1']) # get a cell from a sheet
print(sheet['a1'].value) # get the value from a cell
c = sheet['b1'] # get another cell from the sheet
print(c.value)
# get a row, a column a cell from the sheet
print('row %s, column %s is %s' % (c.row, c.column, c.value))
print('cell %s is %s' % (c.coordinate, c.value))
print(sheet['c1'].value)

print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2): # go through every other row:
    print(i, sheet.cell(row=i, column=2).value)

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(sheet.max_row)  # get the highest row number
print(sheet.max_column) # get the highest column number

# Converting Between Column Letters and Numbers
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
print(get_column_letter(1)) # translate column 1 to a letter
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))

wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(get_column_letter(sheet.max_column))
print(column_index_from_string('A')) # get a's number
print(column_index_from_string('AA'))


import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
print(tuple(sheet['a1':'c3'])) # get all cells from a1 to c3
for rowOfCellObjects in sheet['a1':'c3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('___ END OF ROW ___')

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
print(list(sheet.columns)[1]) # get second column cell
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)

# Writing Excel Documents
# Creating and Saving Excel Documents
import openpyxl
wb = openpyxl.Workbook() # creates a blank workbook
print(wb.sheetnames) # it starts with one sheet
sheet = wb.active
print(sheet.title)
sheet.title = 'Spam bacon eggs' # change the title
print(wb.sheetnames)

import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb.active
sheet.title = 'spam spam'
wb.save('example_copy.xlsx') # save the workbook


# Creating and Removing Sheets
import openpyxl
wb = openpyxl.Workbook()
print(wb.sheetnames)
wb.create_sheet() # add a new sheet
print(wb.sheetnames)
# create a new sheet at index 0
wb.create_sheet(index=0, title='first sheet')
print(wb.sheetnames)
wb.create_sheet(index=2, title='middle sheet')
print(wb.sheetnames)
del wb['middle sheet']
del wb['Sheet1']
print(wb.sheetnames)

# Writing Values to Cells
import openpyxl
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet['A1'] = 'Hello, world!'  # edit the cell's value
print(sheet['A1'].value)

# Setting the Font Style of Cells
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.Workbook()
sheet = wb['Sheet']
italic24Font = Font(size=24, italic=True) # Create a font
sheet['A1'].font = italic24Font # apply the font to a1
sheet['A1'] = 'hello world'
wb.save('styles.xlsx')

# Font Objects
# Formulas
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)' # set the formula
wb.save('writeFormula.xlsx')

# Adjusting Rows and Columns
# Setting Row Height and Column Width

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 'Tall row'
sheet['B2'] = 'Wide Column'
sheet.row_dimensions[1].height = 70
sheet.column_dimensions['B'].width = 20
wb.save('dimensions.xlsx')

# Merging and Unmerging Cells
import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet.merge_cells('A1:D3') # merge all these cells
sheet['A1'] = 'twelve cells merged together'
sheet.merge_cells('C5:D5') # merge these two cells
sheet['C5'] = 'Two merged cells'
wb.save('merged.xlsx')

# Freezing Panes
"""
#  Charts

import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11): # create some data in column A
    sheet['A' + str(i)] = i

refObj = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
seriesObj = openpyxl.chart.Series(refObj, title='First series')
chartObj = openpyxl.chart.PieChart()
chartObj.title = 'My Chart'
chartObj.append(seriesObj)
sheet.add_chart(chartObj, 'C5')
wb.save('sampleChart.xlsx')


























