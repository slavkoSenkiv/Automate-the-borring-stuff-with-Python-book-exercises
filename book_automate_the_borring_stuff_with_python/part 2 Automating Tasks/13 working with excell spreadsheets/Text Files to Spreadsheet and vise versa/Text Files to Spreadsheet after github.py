#! python3
# textFilesToSpreadsheet.py - Reads the content of text file(s),
# and inserts content into new spreadsheet.

import openpyxl, sys

if len(sys.argv) > 1:

    wb = openpyxl.Workbook()
    sheet = wb.active

    files = sys.argv[1:]

    # Each file will be stored as its own column. (x, y) = (c, r)
    for c in range(len(files)):

        text_file = files[c]

        try:
            opened_file = open(text_file)

        except Exception as e:
            print(e)

        lines = opened_file.readlines()

        opened_file.close()

        # First row will be used as a header, using name of text file.
        sheet.cell(row=1, column=c + 1).value = text_file[:-4].upper()

        # Each line will be saved as its own row, after the header.
        for r in range(len(lines)):
            sheet.cell(row=r + 2, column=c + 1).value = lines[r].strip()

        print(text_file + ' was successfully read.')

    # Use the first filename given in saving new filename.
    p = files[0][:-4] + '_text_files.xlsx'

    wb.save(p)

    print('Text files successfully saved to spreadsheet as ' + p)

else:
    print('You must include file name(s) in your argument.')