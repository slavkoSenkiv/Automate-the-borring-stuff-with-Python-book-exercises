import openpyxl

wb = openpyxl.load_workbook('freezeExample.xlsx')
sheet = wb['Sheet']

product_prices = {
    'Garlic': 1.11,
    'Celery': 2.22,
    'Lemon': 3.33}

for rowNum in range(2, sheet.max_row):
    productName = sheet.cell(row=rowNum, column=1).value
    if productName in product_prices:
        sheet.cell(row=rowNum, column=2).value = product_prices[productName]

wb.save('edited_freezeExample_mytryafterbook.xlsx')