import openpyxl

workBookName = 'myProduce'
insertAfter = 4
howManyInsert = 2

inputWb = openpyxl.load_workbook(f'{workBookName}.xlsx')
inputSheet = inputWb.active

outputWb = openpyxl.Workbook()
outputSheet = outputWb.active

for rowNum in range(1, inputSheet.max_row + 1):
    for colNum in range(1, inputSheet.max_column + 1):
        if rowNum <= insertAfter:
            outputSheet.cell(row=rowNum, column=colNum).value = inputSheet.cell(row=rowNum, column=colNum).value
        else:
            outputSheet.cell(row=rowNum + howManyInsert, column=colNum).value = inputSheet.cell(row=rowNum, column=colNum).value

outputWb.save(f'new_{workBookName}_afterGitHub.xlsx')