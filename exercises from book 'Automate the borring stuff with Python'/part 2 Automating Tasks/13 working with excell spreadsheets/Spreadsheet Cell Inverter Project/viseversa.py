import openpyxl

inputWb = openpyxl.load_workbook('testsSheet.xlsx')
inputSheet = inputWb.active

outputWb = openpyxl.Workbook()
outputSheet = outputWb.active

for rowsNum in range(1, inputSheet.max_row + 1):
    for colsNum in range(1, inputSheet.max_column + 1):
        outputSheet.cell(row=colsNum, column=rowsNum).value = inputSheet.cell(row=rowsNum, column=colsNum).value


outputWb.save('new_testSheet.xlsx')