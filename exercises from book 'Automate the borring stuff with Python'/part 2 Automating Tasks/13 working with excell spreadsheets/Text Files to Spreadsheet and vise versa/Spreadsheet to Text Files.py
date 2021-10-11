import openpyxl

wb = openpyxl.load_workbook('testSheet.xlsx')
sheet = wb.active

for column in range(sheet.max_column):
    print('column', column + 1)
    newTxt = open(f'newTxt_{column + 1}.txt', 'w')
    for row in range(sheet.max_row):
        print('row', row + 1)
        if sheet.cell(column=column + 1, row=row + 1).value != None:
            newTxt.write(sheet.cell(column=column + 1, row=row + 1).value)
    newTxt.close()

print('done')



