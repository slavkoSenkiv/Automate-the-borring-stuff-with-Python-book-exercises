#! python3

import openpyxl
from openpyxl.styles import Font
import sys
wb = openpyxl.Workbook()
sheet = wb.active
boldNum = Font(bold=True)

if len(sys.argv) == 2:
    if type(sys.argv[2]) == int:

        userNumber = int(sys.argv[2])

        for firstRow in range(1, userNumber + 1):
            sheet.cell(row=1, column=firstRow + 1).value = firstRow
            sheet.cell(row=1, column=firstRow + 1).font = boldNum
        for firstColumn in range(1, userNumber + 1):
            sheet.cell(row=firstColumn + 1, column=1).value = firstColumn
            sheet.cell(row=firstColumn + 1, column=1).font = boldNum

        for rowCells in range(2, userNumber + 2):
            for colCells in range(2, userNumber + 2):
                sheet.cell(row=rowCells, column=colCells).value = (rowCells - 1) * (colCells - 1)

wb.save('2d_try.xlsx')